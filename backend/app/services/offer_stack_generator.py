"""
Interactive Offer Stack Generator
Dynamic Excel/PowerPoint export with what-if analysis and AI optimization
"""

from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass
from datetime import datetime, timedelta
from decimal import Decimal
import json
import logging
from enum import Enum
import pandas as pd
import numpy as np
from io import BytesIO

from app.services.claude_service import ClaudeService
from app.services.financial_intelligence import FinancialIntelligenceEngine

logger = logging.getLogger(__name__)

class FundingSource(Enum):
    """Types of funding sources"""
    CASH = "cash"
    DEBT = "debt"
    SELLER_FINANCING = "seller_financing"
    EARNOUT = "earnout"
    EQUITY_ROLLOVER = "equity_rollover"
    PREFERRED_EQUITY = "preferred_equity"
    MEZZANINE = "mezzanine"

class DealStructure(Enum):
    """Deal structure types"""
    ASSET_PURCHASE = "asset_purchase"
    STOCK_PURCHASE = "stock_purchase"
    MERGER = "merger"
    MANAGEMENT_BUYOUT = "management_buyout"
    LEVERAGED_BUYOUT = "leveraged_buyout"

@dataclass
class FundingComponent:
    """Individual funding component"""
    source: FundingSource
    amount: Decimal
    percentage: float
    terms: Dict[str, Any]
    cost_of_capital: float
    risk_factor: float

@dataclass
class OfferScenario:
    """Complete offer scenario"""
    scenario_id: str
    scenario_name: str
    total_enterprise_value: Decimal
    purchase_price: Decimal
    funding_components: List[FundingComponent]
    deal_structure: DealStructure
    closing_conditions: List[str]
    timeline: Dict[str, datetime]
    risk_score: float
    confidence_level: float
    ai_insights: str

@dataclass
class WhatIfAnalysis:
    """What-if analysis parameters"""
    base_scenario: OfferScenario
    variable_parameter: str  # e.g., 'purchase_price', 'debt_ratio'
    min_value: float
    max_value: float
    step_size: float
    sensitivity_results: Dict[str, List[float]]

@dataclass
class ExportPackage:
    """Generated export package"""
    excel_file_path: str
    powerpoint_file_path: str
    pdf_summary_path: str
    interactive_dashboard_url: Optional[str]
    generation_timestamp: datetime

class InteractiveOfferStackGenerator:
    """
    Interactive Offer Stack Generator

    Creates sophisticated M&A offer packages with:
    - Multiple funding scenario modeling
    - Real-time what-if analysis
    - AI-powered structure optimization
    - Professional Excel/PowerPoint exports
    - Interactive sensitivity analysis
    - Risk-adjusted returns calculation
    """

    def __init__(self, financial_engine: FinancialIntelligenceEngine):
        self.financial_engine = financial_engine
        self.claude_service = ClaudeService()

    async def generate_offer_stack(
        self,
        target_company_id: str,
        buyer_profile: Dict[str, Any],
        deal_parameters: Dict[str, Any]
    ) -> List[OfferScenario]:
        """
        Generate comprehensive offer stack with multiple scenarios

        Creates 3-5 different scenarios:
        1. Conservative (Low risk, standard terms)
        2. Aggressive (Higher price, more debt)
        3. Creative (Earnouts, seller financing)
        4. Management-friendly (Equity rollover)
        5. Strategic (Synergies-based pricing)
        """

        logger.info(f"Generating offer stack for company {target_company_id}")

        # Step 1: Get financial intelligence
        financial_analysis = await self.financial_engine.analyze_company_financials(
            target_company_id, include_projections=True
        )

        # Step 2: Calculate base valuation range
        base_valuation = await self._calculate_base_valuation(
            financial_analysis, buyer_profile
        )

        # Step 3: Generate multiple scenarios
        scenarios = []

        # Conservative scenario
        conservative = await self._generate_conservative_scenario(
            base_valuation, financial_analysis, deal_parameters
        )
        scenarios.append(conservative)

        # Aggressive scenario
        aggressive = await self._generate_aggressive_scenario(
            base_valuation, financial_analysis, deal_parameters
        )
        scenarios.append(aggressive)

        # Creative scenario
        creative = await self._generate_creative_scenario(
            base_valuation, financial_analysis, deal_parameters
        )
        scenarios.append(creative)

        # Management-friendly scenario (if applicable)
        if deal_parameters.get('management_participation', False):
            mgmt_friendly = await self._generate_management_friendly_scenario(
                base_valuation, financial_analysis, deal_parameters
            )
            scenarios.append(mgmt_friendly)

        # Strategic scenario (if synergies identified)
        if buyer_profile.get('synergies_potential', 0) > 0:
            strategic = await self._generate_strategic_scenario(
                base_valuation, financial_analysis, deal_parameters, buyer_profile
            )
            scenarios.append(strategic)

        # Step 4: Apply AI optimization
        optimized_scenarios = []
        for scenario in scenarios:
            optimized = await self._optimize_scenario_with_ai(
                scenario, financial_analysis, buyer_profile
            )
            optimized_scenarios.append(optimized)

        return optimized_scenarios

    async def perform_what_if_analysis(
        self,
        base_scenario: OfferScenario,
        analysis_parameters: WhatIfAnalysis
    ) -> WhatIfAnalysis:
        """
        Perform comprehensive what-if sensitivity analysis

        Analyzes impact of changing key variables:
        - Purchase price variations
        - Debt/equity mix changes
        - Interest rate sensitivity
        - Earnout percentage impact
        - Timeline adjustments
        """

        logger.info(f"Performing what-if analysis for scenario {base_scenario.scenario_id}")

        results = {}

        # Generate value range for the parameter
        param_values = np.arange(
            analysis_parameters.min_value,
            analysis_parameters.max_value + analysis_parameters.step_size,
            analysis_parameters.step_size
        )

        # Calculate impact on key metrics
        for param_value in param_values:
            # Create modified scenario
            modified_scenario = await self._modify_scenario_parameter(
                base_scenario,
                analysis_parameters.variable_parameter,
                param_value
            )

            # Calculate impact metrics
            impact_metrics = await self._calculate_scenario_metrics(modified_scenario)

            # Store results
            for metric, value in impact_metrics.items():
                if metric not in results:
                    results[metric] = []
                results[metric].append(value)

        analysis_parameters.sensitivity_results = results
        return analysis_parameters

    async def export_offer_package(
        self,
        scenarios: List[OfferScenario],
        target_company_id: str,
        export_options: Dict[str, Any]
    ) -> ExportPackage:
        """
        Export comprehensive offer package to Excel, PowerPoint, and PDF

        Excel Features:
        - Interactive scenario comparison dashboard
        - Dynamic charts and graphs
        - What-if analysis tables
        - Financial projections
        - Sensitivity analysis charts

        PowerPoint Features:
        - Executive summary slides
        - Deal structure visualization
        - Financing sources breakdown
        - Risk/return analysis
        - Implementation timeline
        """

        logger.info(f"Exporting offer package for {len(scenarios)} scenarios")

        timestamp = datetime.utcnow()
        export_prefix = f"offer_package_{target_company_id}_{timestamp.strftime('%Y%m%d_%H%M%S')}"

        # Generate Excel workbook
        excel_path = await self._generate_excel_workbook(
            scenarios, target_company_id, export_prefix, export_options
        )

        # Generate PowerPoint presentation
        ppt_path = await self._generate_powerpoint_presentation(
            scenarios, target_company_id, export_prefix, export_options
        )

        # Generate PDF summary
        pdf_path = await self._generate_pdf_summary(
            scenarios, target_company_id, export_prefix, export_options
        )

        # Generate interactive dashboard (optional)
        dashboard_url = None
        if export_options.get('include_dashboard', False):
            dashboard_url = await self._generate_interactive_dashboard(
                scenarios, target_company_id
            )

        return ExportPackage(
            excel_file_path=excel_path,
            powerpoint_file_path=ppt_path,
            pdf_summary_path=pdf_path,
            interactive_dashboard_url=dashboard_url,
            generation_timestamp=timestamp
        )

    # Private methods for scenario generation

    async def _calculate_base_valuation(
        self,
        financial_analysis,
        buyer_profile: Dict[str, Any]
    ) -> Tuple[Decimal, Decimal]:
        """Calculate base valuation range"""

        # Use financial intelligence valuation as starting point
        low_val, high_val = financial_analysis.valuation_range

        # Adjust for buyer-specific factors
        buyer_multiple = buyer_profile.get('typical_multiple', 1.0)
        synergies_value = buyer_profile.get('synergies_value', 0)

        adjusted_low = Decimal(low_val * buyer_multiple)
        adjusted_high = Decimal(high_val * buyer_multiple + synergies_value)

        return adjusted_low, adjusted_high

    async def _generate_conservative_scenario(
        self,
        base_valuation: Tuple[Decimal, Decimal],
        financial_analysis,
        deal_parameters: Dict[str, Any]
    ) -> OfferScenario:
        """Generate conservative scenario"""

        low_val, high_val = base_valuation
        purchase_price = low_val + (high_val - low_val) * Decimal('0.3')  # 30% of range

        # Conservative funding structure: 70% cash, 30% debt
        funding_components = [
            FundingComponent(
                source=FundingSource.CASH,
                amount=purchase_price * Decimal('0.7'),
                percentage=70.0,
                terms={'immediate': True},
                cost_of_capital=0.05,
                risk_factor=0.1
            ),
            FundingComponent(
                source=FundingSource.DEBT,
                amount=purchase_price * Decimal('0.3'),
                percentage=30.0,
                terms={'term': '5 years', 'rate': '6%'},
                cost_of_capital=0.06,
                risk_factor=0.3
            )
        ]

        return OfferScenario(
            scenario_id=f"conservative_{datetime.utcnow().timestamp()}",
            scenario_name="Conservative Offer",
            total_enterprise_value=purchase_price,
            purchase_price=purchase_price,
            funding_components=funding_components,
            deal_structure=DealStructure.STOCK_PURCHASE,
            closing_conditions=self._get_standard_closing_conditions(),
            timeline=self._get_conservative_timeline(),
            risk_score=0.3,  # Low risk
            confidence_level=0.9,  # High confidence
            ai_insights=""
        )

    async def _generate_aggressive_scenario(
        self,
        base_valuation: Tuple[Decimal, Decimal],
        financial_analysis,
        deal_parameters: Dict[str, Any]
    ) -> OfferScenario:
        """Generate aggressive scenario"""

        low_val, high_val = base_valuation
        purchase_price = low_val + (high_val - low_val) * Decimal('0.9')  # 90% of range

        # Aggressive funding: 40% cash, 60% debt
        funding_components = [
            FundingComponent(
                source=FundingSource.CASH,
                amount=purchase_price * Decimal('0.4'),
                percentage=40.0,
                terms={'immediate': True},
                cost_of_capital=0.05,
                risk_factor=0.1
            ),
            FundingComponent(
                source=FundingSource.DEBT,
                amount=purchase_price * Decimal('0.5'),
                percentage=50.0,
                terms={'term': '7 years', 'rate': '7.5%'},
                cost_of_capital=0.075,
                risk_factor=0.6
            ),
            FundingComponent(
                source=FundingSource.MEZZANINE,
                amount=purchase_price * Decimal('0.1'),
                percentage=10.0,
                terms={'term': '5 years', 'rate': '12%'},
                cost_of_capital=0.12,
                risk_factor=0.7
            )
        ]

        return OfferScenario(
            scenario_id=f"aggressive_{datetime.utcnow().timestamp()}",
            scenario_name="Aggressive Offer",
            total_enterprise_value=purchase_price,
            purchase_price=purchase_price,
            funding_components=funding_components,
            deal_structure=DealStructure.LEVERAGED_BUYOUT,
            closing_conditions=self._get_standard_closing_conditions(),
            timeline=self._get_fast_timeline(),
            risk_score=0.8,  # High risk
            confidence_level=0.6,  # Lower confidence
            ai_insights=""
        )

    async def _generate_creative_scenario(
        self,
        base_valuation: Tuple[Decimal, Decimal],
        financial_analysis,
        deal_parameters: Dict[str, Any]
    ) -> OfferScenario:
        """Generate creative scenario with earnouts and seller financing"""

        low_val, high_val = base_valuation
        base_price = low_val + (high_val - low_val) * Decimal('0.6')  # 60% of range
        earnout_potential = (high_val - base_price) * Decimal('0.8')  # 80% of upside

        funding_components = [
            FundingComponent(
                source=FundingSource.CASH,
                amount=base_price * Decimal('0.5'),
                percentage=35.0,  # % of total potential value
                terms={'immediate': True},
                cost_of_capital=0.05,
                risk_factor=0.1
            ),
            FundingComponent(
                source=FundingSource.SELLER_FINANCING,
                amount=base_price * Decimal('0.3'),
                percentage=21.0,
                terms={'term': '4 years', 'rate': '5%', 'monthly_payments': True},
                cost_of_capital=0.05,
                risk_factor=0.2
            ),
            FundingComponent(
                source=FundingSource.DEBT,
                amount=base_price * Decimal('0.2'),
                percentage=14.0,
                terms={'term': '5 years', 'rate': '6.5%'},
                cost_of_capital=0.065,
                risk_factor=0.4
            ),
            FundingComponent(
                source=FundingSource.EARNOUT,
                amount=earnout_potential,
                percentage=30.0,
                terms={
                    'performance_period': '3 years',
                    'metrics': ['revenue_growth', 'ebitda_margin'],
                    'thresholds': {'revenue_growth': 0.15, 'ebitda_margin': 0.20}
                },
                cost_of_capital=0.15,  # Higher due to uncertainty
                risk_factor=0.9
            )
        ]

        total_value = base_price + earnout_potential

        return OfferScenario(
            scenario_id=f"creative_{datetime.utcnow().timestamp()}",
            scenario_name="Creative Structure",
            total_enterprise_value=total_value,
            purchase_price=base_price,  # Base purchase price
            funding_components=funding_components,
            deal_structure=DealStructure.ASSET_PURCHASE,
            closing_conditions=self._get_earnout_closing_conditions(),
            timeline=self._get_standard_timeline(),
            risk_score=0.6,  # Moderate risk
            confidence_level=0.7,
            ai_insights=""
        )

    async def _optimize_scenario_with_ai(
        self,
        scenario: OfferScenario,
        financial_analysis,
        buyer_profile: Dict[str, Any]
    ) -> OfferScenario:
        """Apply AI optimization to scenario"""

        optimization_prompt = f"""
        Optimize this M&A offer scenario based on financial analysis and buyer profile:

        SCENARIO: {scenario.scenario_name}
        Purchase Price: ${scenario.purchase_price:,.0f}
        Enterprise Value: ${scenario.total_enterprise_value:,.0f}

        FUNDING STRUCTURE:
        {self._format_funding_for_ai(scenario.funding_components)}

        TARGET COMPANY METRICS:
        Revenue Growth: {financial_analysis.key_metrics.get('revenue_growth', 0):.1%}
        EBITDA Margin: {financial_analysis.key_metrics.get('ebitda_margin', 0):.1%}
        Debt-to-Equity: {financial_analysis.key_metrics.get('debt_to_equity', 0):.2f}
        Risk Score: {financial_analysis.deal_readiness_score}/100

        BUYER PROFILE:
        Cost of Capital: {buyer_profile.get('wacc', 0.10):.1%}
        Risk Tolerance: {buyer_profile.get('risk_tolerance', 'moderate')}
        Strategic Synergies: ${buyer_profile.get('synergies_value', 0):,.0f}

        Please provide optimization recommendations for:
        1. Funding structure improvements
        2. Risk mitigation strategies
        3. Terms optimization
        4. Timeline adjustments
        5. Closing conditions refinements

        Return specific, actionable recommendations.
        """

        ai_insights = await self.claude_service.analyze_content(optimization_prompt)
        scenario.ai_insights = ai_insights

        # Apply AI recommendations (simplified implementation)
        # In a full implementation, this would parse AI recommendations and modify the scenario

        return scenario

    # Export generation methods

    async def _generate_excel_workbook(
        self,
        scenarios: List[OfferScenario],
        target_company_id: str,
        export_prefix: str,
        export_options: Dict[str, Any]
    ) -> str:
        """Generate comprehensive Excel workbook"""

        from openpyxl import Workbook
        from openpyxl.chart import LineChart, Reference, BarChart
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create sheets
        summary_sheet = wb.create_sheet("Executive Summary")
        comparison_sheet = wb.create_sheet("Scenario Comparison")
        funding_sheet = wb.create_sheet("Funding Analysis")
        sensitivity_sheet = wb.create_sheet("Sensitivity Analysis")
        timeline_sheet = wb.create_sheet("Implementation Timeline")

        # Populate Executive Summary
        await self._populate_excel_summary(summary_sheet, scenarios, target_company_id)

        # Populate Scenario Comparison
        await self._populate_excel_comparison(comparison_sheet, scenarios)

        # Populate Funding Analysis
        await self._populate_excel_funding(funding_sheet, scenarios)

        # Add charts and formatting
        await self._add_excel_charts(wb, scenarios)

        # Save workbook
        excel_path = f"exports/{export_prefix}.xlsx"
        wb.save(excel_path)

        return excel_path

    async def _generate_powerpoint_presentation(
        self,
        scenarios: List[OfferScenario],
        target_company_id: str,
        export_prefix: str,
        export_options: Dict[str, Any]
    ) -> str:
        """Generate PowerPoint presentation"""

        from pptx import Presentation
        from pptx.util import Inches

        prs = Presentation()

        # Title slide
        await self._add_ppt_title_slide(prs, target_company_id, scenarios)

        # Executive summary
        await self._add_ppt_executive_summary(prs, scenarios)

        # Individual scenario slides
        for scenario in scenarios:
            await self._add_ppt_scenario_slide(prs, scenario)

        # Comparison slide
        await self._add_ppt_comparison_slide(prs, scenarios)

        # Next steps slide
        await self._add_ppt_next_steps_slide(prs, scenarios)

        # Save presentation
        ppt_path = f"exports/{export_prefix}.pptx"
        prs.save(ppt_path)

        return ppt_path

    # Helper methods

    def _get_standard_closing_conditions(self) -> List[str]:
        """Get standard closing conditions"""
        return [
            "Completion of satisfactory due diligence",
            "No material adverse change",
            "Regulatory approvals obtained",
            "Third-party consents received",
            "Financing commitments secured",
            "Key employee retention agreements"
        ]

    def _get_earnout_closing_conditions(self) -> List[str]:
        """Get closing conditions for earnout scenarios"""
        conditions = self._get_standard_closing_conditions()
        conditions.extend([
            "Earnout calculation mechanism agreed",
            "Performance monitoring systems established",
            "Dispute resolution procedures defined"
        ])
        return conditions

    def _get_conservative_timeline(self) -> Dict[str, datetime]:
        """Get conservative transaction timeline"""
        now = datetime.utcnow()
        return {
            'loi_signed': now + timedelta(days=14),
            'due_diligence_complete': now + timedelta(days=60),
            'definitive_agreement': now + timedelta(days=90),
            'closing': now + timedelta(days=120)
        }

    def _get_fast_timeline(self) -> Dict[str, datetime]:
        """Get aggressive transaction timeline"""
        now = datetime.utcnow()
        return {
            'loi_signed': now + timedelta(days=7),
            'due_diligence_complete': now + timedelta(days=35),
            'definitive_agreement': now + timedelta(days=60),
            'closing': now + timedelta(days=75)
        }

    def _get_standard_timeline(self) -> Dict[str, datetime]:
        """Get standard transaction timeline"""
        now = datetime.utcnow()
        return {
            'loi_signed': now + timedelta(days=10),
            'due_diligence_complete': now + timedelta(days=45),
            'definitive_agreement': now + timedelta(days=75),
            'closing': now + timedelta(days=105)
        }

    def _format_funding_for_ai(self, components: List[FundingComponent]) -> str:
        """Format funding components for AI analysis"""
        formatted = []
        for comp in components:
            formatted.append(
                f"- {comp.source.value}: ${comp.amount:,.0f} ({comp.percentage:.1f}%) "
                f"at {comp.cost_of_capital:.1%} cost"
            )
        return "\n".join(formatted)

    async def _modify_scenario_parameter(
        self,
        scenario: OfferScenario,
        parameter: str,
        new_value: float
    ) -> OfferScenario:
        """Create modified scenario with changed parameter"""
        # Create a copy and modify the specified parameter
        # Implementation would depend on parameter type
        return scenario  # Placeholder

    async def _calculate_scenario_metrics(self, scenario: OfferScenario) -> Dict[str, float]:
        """Calculate key metrics for scenario"""
        return {
            'irr': 0.25,  # Placeholder calculations
            'cash_multiple': 2.5,
            'payback_period': 3.2,
            'debt_service_coverage': 1.8
        }

    # Placeholder methods for Excel/PowerPoint population
    async def _populate_excel_summary(self, sheet, scenarios, company_id): pass
    async def _populate_excel_comparison(self, sheet, scenarios): pass
    async def _populate_excel_funding(self, sheet, scenarios): pass
    async def _add_excel_charts(self, workbook, scenarios): pass

    async def _add_ppt_title_slide(self, prs, company_id, scenarios): pass
    async def _add_ppt_executive_summary(self, prs, scenarios): pass
    async def _add_ppt_scenario_slide(self, prs, scenario): pass
    async def _add_ppt_comparison_slide(self, prs, scenarios): pass
    async def _add_ppt_next_steps_slide(self, prs, scenarios): pass

    async def _generate_pdf_summary(self, scenarios, company_id, prefix, options) -> str:
        return f"exports/{prefix}.pdf"  # Placeholder

    async def _generate_interactive_dashboard(self, scenarios, company_id) -> str:
        return f"https://dashboard.example.com/offers/{company_id}"  # Placeholder